from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from auth import get_current_user
from database import get_db
from models import Board
import io

router = APIRouter(prefix="/export", tags=["export"])
db = get_db()


@router.get("/excel/{board_id}")
async def export_board_to_excel(
    board_id: str,
    current_user: dict = Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    board = await db.boards.find_one({"id": board_id})
    if not board:
        raise HTTPException(status_code=404, detail="Board not found")

    groups = await db.groups.find({"board_id": board_id}).sort("position", 1).to_list(1000)
    items = await db.items.find({"board_id": board_id}).to_list(10000)

    columns = board.get("columns", [])
    if not columns:
        raise HTTPException(status_code=400, detail="Board has no columns")

    # Build column map: id -> (title, index, options_map)
    col_info = []
    for col in columns:
        options_map = {}
        for opt in col.get("options", []):
            options_map[opt.get("id", "")] = opt.get("label", "")
        col_info.append({
            "id": col["id"],
            "title": col["title"],
            "type": col.get("type", "text"),
            "options_map": options_map
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = board.get("name", "Board")[:31]

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    group_font = Font(bold=True, size=12, color="333333")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    group_colors = [
        "579BFC", "A25DDC", "00C875", "FDAB3D", "E2445C",
        "0086C0", "037F4C", "CAA2FF", "FF5AC4", "66CCFF"
    ]

    # Build items by group_id
    items_by_group = {}
    ungrouped_items = []
    for item in items:
        gid = item.get("group_id")
        if gid:
            items_by_group.setdefault(gid, []).append(item)
        else:
            ungrouped_items.append(item)

    current_row = 1

    def write_header_row(row_num):
        for col_idx, ci in enumerate(col_info, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=ci["title"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        return row_num + 1

    def write_items(item_list, start_row):
        row = start_row
        for item in item_list:
            for col_idx, ci in enumerate(col_info, 1):
                if col_idx == 1:
                    # First column is always the item name
                    ws.cell(row=row, column=col_idx, value=item.get("name", ""))
                else:
                    raw_val = item.get("column_values", {}).get(ci["id"])
                    if raw_val is None:
                        ws.cell(row=row, column=col_idx, value="")
                    elif isinstance(raw_val, dict):
                        ws.cell(row=row, column=col_idx, value=raw_val.get("label", raw_val.get("text", str(raw_val))))
                    else:
                        # For status/priority, try to resolve option label
                        if ci["type"] in ("status", "priority") and ci["options_map"]:
                            label = ci["options_map"].get(str(raw_val), str(raw_val))
                            ws.cell(row=row, column=col_idx, value=label)
                        elif ci["type"] == "link" and str(raw_val).startswith("http"):
                            cell = ws.cell(row=row, column=col_idx, value=str(raw_val))
                            cell.hyperlink = str(raw_val)
                            cell.font = Font(color="0563C1", underline="single")
                        else:
                            ws.cell(row=row, column=col_idx, value=str(raw_val))
                ws.cell(row=row, column=col_idx).border = thin_border
            row += 1
        return row

    for g_idx, group in enumerate(groups):
        group_id = group["id"]
        group_items = items_by_group.get(group_id, [])

        # Write group name row
        cell = ws.cell(row=current_row, column=1, value=group.get("title", "Group"))
        color_hex = group_colors[g_idx % len(group_colors)]
        cell.font = Font(bold=True, size=12, color=color_hex)
        current_row += 1

        # Write header row
        current_row = write_header_row(current_row)

        # Write items
        current_row = write_items(group_items, current_row)

        # Blank row separator
        current_row += 1

    # Handle ungrouped items
    if ungrouped_items:
        cell = ws.cell(row=current_row, column=1, value="Other Items")
        cell.font = group_font
        current_row += 1
        current_row = write_header_row(current_row)
        current_row = write_items(ungrouped_items, current_row)

    # Auto-adjust column widths
    for col_idx in range(1, len(col_info) + 1):
        max_len = 10
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_val in row_cells:
                if cell_val:
                    max_len = max(max_len, min(len(str(cell_val)), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 4

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = board.get("name", "board").replace(" ", "_")[:50]
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'}
    )

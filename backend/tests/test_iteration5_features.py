"""
Iteration 5 feature tests:
1) Enhanced Excel Import with group detection from blank-row separators
2) Excel Import with existing_board_id (append mode)
3) GET /api/export/excel/{board_id} - Export to .xlsx
4) Chart configs persistence via PUT/GET /api/boards/{id}
"""
import os
import io
import uuid
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://acuity-team-hub.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

TEST_EMAIL = "testuser@acuity.com"
TEST_PASS = "TestPass123!"
TEST_XLSX = "/tmp/test_pipeline.xlsx"


# ---------- Fixtures ----------
@pytest.fixture(scope="module")
def auth_headers():
    s = requests.Session()
    # Try login, fall back to register
    r = s.post(f"{API}/auth/login", json={"email": TEST_EMAIL, "password": TEST_PASS}, timeout=30)
    if r.status_code != 200:
        s.post(f"{API}/auth/register", json={
            "email": TEST_EMAIL, "password": TEST_PASS, "name": "Test User"
        }, timeout=30)
        r = s.post(f"{API}/auth/login", json={"email": TEST_EMAIL, "password": TEST_PASS}, timeout=30)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    token = r.json().get("access_token") or r.json().get("token")
    assert token
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture(scope="module")
def workspace_id(auth_headers):
    r = requests.get(f"{API}/workspaces", headers=auth_headers, timeout=30)
    assert r.status_code == 200, r.text
    wss = r.json()
    assert len(wss) > 0, "No workspace available"
    return wss[0]["id"]


# ---------- 1) Enhanced Import with group detection ----------
class TestImportGroupDetection:
    board_id = None

    def test_01_file_exists(self):
        assert os.path.exists(TEST_XLSX), f"Test file missing: {TEST_XLSX}"
        assert os.path.getsize(TEST_XLSX) > 1000

    def test_02_import_creates_multiple_groups(self, auth_headers, workspace_id):
        with open(TEST_XLSX, "rb") as f:
            files = {"file": ("Pipeline.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = requests.post(
                f"{API}/import/excel?workspace_id={workspace_id}&board_name=TEST_Pipeline_Import",
                headers=auth_headers,
                files=files,
                timeout=90,
            )
        assert r.status_code == 200, f"Import failed: {r.status_code} {r.text}"
        data = r.json()
        assert "board_id" in data
        assert "groups_created" in data
        assert "items_created" in data
        TestImportGroupDetection.board_id = data["board_id"]
        groups = data["groups_created"]
        # Expected groups: Pipeline, Completed, NFA, Enquiry - Working On
        print(f"Groups created: {groups}, items: {data['items_created']}")
        assert len(groups) >= 3, f"Expected >=3 groups, got: {groups}"
        # Check key expected group names present (case-insensitive)
        gset = {g.lower() for g in groups}
        expected_any = {"completed", "nfa"}
        assert expected_any & gset, f"Missing expected groups. Got: {groups}"
        assert data["items_created"] > 50, f"Expected many items, got {data['items_created']}"

    def test_03_board_persisted_with_groups(self, auth_headers):
        assert TestImportGroupDetection.board_id is not None
        bid = TestImportGroupDetection.board_id
        rb = requests.get(f"{API}/boards/{bid}", headers=auth_headers, timeout=30)
        assert rb.status_code == 200, rb.text
        # groups list
        rg = requests.get(f"{API}/groups/board/{bid}", headers=auth_headers, timeout=30)
        assert rg.status_code == 200, rg.text
        groups = rg.json()
        assert len(groups) >= 3, f"Expected >=3 groups persisted, got {len(groups)}"
        # items
        ri = requests.get(f"{API}/items/board/{bid}", headers=auth_headers, timeout=30)
        assert ri.status_code == 200, ri.text
        assert len(ri.json()) > 50


# ---------- 2) Import append mode (existing_board_id) ----------
class TestImportAppendMode:
    def test_01_append_to_existing_board(self, auth_headers, workspace_id):
        existing_id = TestImportGroupDetection.board_id
        assert existing_id, "Requires previous import to run first"

        # Record current items/groups count
        pre_groups = requests.get(f"{API}/groups/board/{existing_id}", headers=auth_headers, timeout=30).json()
        pre_items = requests.get(f"{API}/items/board/{existing_id}", headers=auth_headers, timeout=30).json()
        pre_g = len(pre_groups)
        pre_i = len(pre_items)

        with open(TEST_XLSX, "rb") as f:
            files = {"file": ("Pipeline.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = requests.post(
                f"{API}/import/excel?workspace_id={workspace_id}&existing_board_id={existing_id}",
                headers=auth_headers,
                files=files,
                timeout=90,
            )
        assert r.status_code == 200, f"Append import failed: {r.status_code} {r.text}"
        data = r.json()
        assert data["board_id"] == existing_id, "Should append to same board"
        assert data["items_created"] > 0

        # Verify counts grew
        post_groups = requests.get(f"{API}/groups/board/{existing_id}", headers=auth_headers, timeout=30).json()
        post_items = requests.get(f"{API}/items/board/{existing_id}", headers=auth_headers, timeout=30).json()
        assert len(post_groups) > pre_g, "Groups should increase after append"
        assert len(post_items) > pre_i, "Items should increase after append"

    def test_02_missing_workspace_and_board_returns_400(self, auth_headers):
        with open(TEST_XLSX, "rb") as f:
            files = {"file": ("Pipeline.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = requests.post(f"{API}/import/excel", headers=auth_headers, files=files, timeout=60)
        assert r.status_code == 400


# ---------- 3) Export to Excel ----------
class TestExportExcel:
    def test_01_export_returns_xlsx(self, auth_headers):
        bid = TestImportGroupDetection.board_id
        assert bid, "Need a board from import test"
        r = requests.get(f"{API}/export/excel/{bid}", headers=auth_headers, timeout=60)
        assert r.status_code == 200, f"Export failed: {r.status_code} {r.text[:500]}"
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "octet-stream" in ct, f"Unexpected content-type: {ct}"
        # Validate xlsx signature (PK header)
        assert r.content[:2] == b"PK", "Response is not a valid .xlsx (missing ZIP magic)"
        assert len(r.content) > 2000, f"Export too small: {len(r.content)}"
        # Parse workbook and verify it has groups/header content
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        ws = wb.active
        cells = [c.value for row in ws.iter_rows(max_row=50, values_only=True) for c in []]
        # simpler: flatten all values
        vals = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    vals.append(str(v))
        joined = " | ".join(vals[:500])
        print(f"First export content snippet: {joined[:300]}")
        assert len(vals) > 10

    def test_02_export_nonexistent_board_returns_404(self, auth_headers):
        fake_id = str(uuid.uuid4())
        r = requests.get(f"{API}/export/excel/{fake_id}", headers=auth_headers, timeout=30)
        assert r.status_code == 404


# ---------- 4) Chart configs persistence ----------
class TestChartConfigs:
    def test_01_put_chart_configs(self, auth_headers):
        bid = TestImportGroupDetection.board_id
        assert bid
        configs = [
            {"id": "chart-1", "column_id": "col-status", "chart_type": "bar", "title": "Status Distribution"},
            {"id": "chart-2", "column_id": "col-priority", "chart_type": "pie", "title": "Priority Split"},
        ]
        r = requests.put(
            f"{API}/boards/{bid}",
            headers={**auth_headers, "Content-Type": "application/json"},
            json={"chart_configs": configs},
            timeout=30,
        )
        assert r.status_code == 200, f"{r.status_code} {r.text}"
        body = r.json()
        assert "chart_configs" in body
        assert len(body["chart_configs"]) == 2

    def test_02_get_returns_chart_configs(self, auth_headers):
        bid = TestImportGroupDetection.board_id
        r = requests.get(f"{API}/boards/{bid}", headers=auth_headers, timeout=30)
        assert r.status_code == 200
        body = r.json()
        assert "chart_configs" in body
        assert len(body["chart_configs"]) == 2
        titles = {c.get("title") for c in body["chart_configs"]}
        assert "Status Distribution" in titles
        assert "Priority Split" in titles

    def test_03_chart_configs_update_replaces(self, auth_headers):
        bid = TestImportGroupDetection.board_id
        new_configs = [{"id": "chart-x", "column_id": "col-x", "chart_type": "line", "title": "Only One"}]
        r = requests.put(
            f"{API}/boards/{bid}",
            headers={**auth_headers, "Content-Type": "application/json"},
            json={"chart_configs": new_configs},
            timeout=30,
        )
        assert r.status_code == 200
        g = requests.get(f"{API}/boards/{bid}", headers=auth_headers, timeout=30).json()
        assert len(g["chart_configs"]) == 1
        assert g["chart_configs"][0]["title"] == "Only One"


# ---------- Cleanup ----------
class TestCleanup:
    def test_delete_test_board(self, auth_headers):
        bid = TestImportGroupDetection.board_id
        if not bid:
            pytest.skip("No board to cleanup")
        r = requests.delete(f"{API}/boards/{bid}", headers=auth_headers, timeout=30)
        assert r.status_code in (200, 204)
